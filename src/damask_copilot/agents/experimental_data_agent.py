"""Deprecated experimental-data micro-agent retained for compatibility wrappers."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import yaml

from damask_copilot.agents._deprecation import warn_legacy_agent
from damask_copilot.graph.materials_research_state import MaterialsResearchState, append_trace
from damask_copilot.llm.prompts import load_prompt
from damask_copilot.llm.structured_runner import StructuredLLMRunner
from damask_copilot.schemas.llm_outputs import ExperimentalDataInterpretationOutput


class ExperimentalDataAgent:
    """Deprecated wrapper for experimental-data summarization.

    The unified v1 architecture uses `ScientificKnowledgeAgent` instead.
    """

    name = "experimental_data_agent"

    def __init__(
        self,
        *,
        use_llm: bool = False,
        model_name: str | None = None,
        llm_runner: StructuredLLMRunner | None = None,
    ) -> None:
        warn_legacy_agent(legacy_name="ExperimentalDataAgent", replacement="ScientificKnowledgeAgent")
        self.use_llm = use_llm
        self.model_name = model_name
        self.llm_runner = llm_runner

    def run(self, state: MaterialsResearchState) -> MaterialsResearchState:
        summary = self._run_deterministic(state)
        if self.use_llm or state.get("use_llm", False):
            return self._run_llm(state, summary)
        return summary

    def _run_deterministic(self, state: MaterialsResearchState) -> MaterialsResearchState:
        experimental_files = list(state.get("experimental_files", [])) or self._infer_experimental_files(state)
        if not experimental_files:
            updated = dict(state)
            updated["experimental_data_summary"] = {
                "status": "experimental_data_missing",
                "summary": "No experimental datasets were supplied. This is acceptable for exploratory planning, but it prevents quantitative validation.",
                "datasets": [],
                "observable_candidates": [],
                "critical_metadata_missing": False,
                "needs_human_correction": False,
                "semantic_column_guesses": {},
                "metadata_questions": [],
                "interpretation_summary": "No experimental datasets were supplied; experiment-driven validation was not requested by the current state.",
            }
            return append_trace(updated, self.name, "experimental_data_missing", {})

        datasets: list[dict[str, Any]] = []
        observable_candidates: list[str] = []
        critical_metadata_missing = False

        for file_path in experimental_files:
            dataset = self._summarize_file(Path(file_path))
            datasets.append(dataset)
            for item in dataset.get("observable_candidates", []):
                if item not in observable_candidates:
                    observable_candidates.append(item)
            if dataset.get("critical_metadata_missing", False):
                critical_metadata_missing = True

        updated = dict(state)
        updated["experimental_files"] = experimental_files
        updated["experimental_data_summary"] = {
            "status": "experimental_data_loaded",
            "summary": self._build_summary(datasets, critical_metadata_missing),
            "datasets": datasets,
            "observable_candidates": observable_candidates,
            "critical_metadata_missing": critical_metadata_missing,
            "needs_human_correction": critical_metadata_missing,
            "semantic_column_guesses": self._guess_semantic_columns(datasets),
            "metadata_questions": self._build_metadata_questions(datasets),
            "interpretation_summary": "Deterministic experimental-data summary completed.",
        }
        return append_trace(updated, self.name, "experimental_data_summarized", {
            "file_count": len(datasets),
            "critical_metadata_missing": critical_metadata_missing,
        })

    def _run_llm(self, original_state: MaterialsResearchState, summarized_state: MaterialsResearchState) -> MaterialsResearchState:
        summary = dict(summarized_state.get("experimental_data_summary") or {})
        if summary.get("status") == "experimental_data_missing":
            return summarized_state

        runner = self.llm_runner or StructuredLLMRunner(model_name=original_state.get("model") or self.model_name)
        parsed = runner.run_structured(
            prompt_name="experimental_data_agent",
            system_prompt=load_prompt("experimental_data_agent"),
            user_prompt=(
                f"User query: {original_state.get('user_query')}\n"
                f"Research case: {original_state.get('research_case')}\n"
                f"Deterministic experimental summary: {json.dumps(summary, ensure_ascii=False)}"
            ),
            output_schema=ExperimentalDataInterpretationOutput,
            model_name=original_state.get("model") or self.model_name,
        )

        updated = dict(summarized_state)
        llm_guesses = dict(parsed.semantic_column_guesses)
        combined_guesses = dict(summary.get("semantic_column_guesses", {}))
        combined_guesses.update({key: value for key, value in llm_guesses.items() if value})
        observable_candidates = list(summary.get("observable_candidates", []))
        for item in parsed.likely_observables:
            if item not in observable_candidates:
                observable_candidates.append(item)
        metadata_questions = list(summary.get("metadata_questions", []))
        for item in parsed.metadata_questions:
            if item not in metadata_questions:
                metadata_questions.append(item)
        summary["semantic_column_guesses"] = combined_guesses
        summary["observable_candidates"] = observable_candidates
        summary["metadata_questions"] = metadata_questions
        summary["interpretation_summary"] = parsed.interpretation_summary or summary.get("summary", "")
        summary["needs_human_correction"] = bool(summary.get("critical_metadata_missing") or metadata_questions)
        summary["llm_interpretation"] = parsed.model_dump()
        updated["experimental_data_summary"] = summary
        return append_trace(updated, self.name, "experimental_data_interpreted_llm", {
            "observable_candidates": observable_candidates,
            "metadata_question_count": len(metadata_questions),
        })

    def _infer_experimental_files(self, state: MaterialsResearchState) -> list[str]:
        supported = {".csv", ".xlsx", ".xls", ".yaml", ".yml", ".json", ".txt"}
        return [
            path
            for path in state.get("user_files", [])
            if Path(path).suffix.lower() in supported
        ]

    def _summarize_file(self, path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._summarize_csv(path)
        if suffix in {".yaml", ".yml"}:
            return self._summarize_yaml(path)
        if suffix == ".json":
            return self._summarize_json(path)
        if suffix == ".txt":
            return self._summarize_txt(path)
        if suffix in {".xlsx", ".xls"}:
            return self._summarize_xlsx(path)
        return {
            "path": str(path),
            "format": suffix.lstrip(".") or "unknown",
            "rows": None,
            "columns": [],
            "units_detected": [],
            "missing_values": None,
            "observable_candidates": [],
            "critical_metadata_missing": True,
            "warnings": [f"Unsupported experimental file format: {suffix or 'unknown'}"],
        }

    def _summarize_csv(self, path: Path) -> dict[str, Any]:
        with path.open("r", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
            columns = list(reader.fieldnames or [])
        return self._tabular_summary(path, "csv", rows, columns)

    def _summarize_xlsx(self, path: Path) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "path": str(path),
                "format": "xlsx",
                "rows": None,
                "columns": [],
                "units_detected": [],
                "missing_values": None,
                "observable_candidates": [],
                "critical_metadata_missing": True,
                "warnings": ["openpyxl is not available; XLSX metadata could not be inspected."],
            }
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(item) if item is not None else "" for item in (values[0] if values else [])]
        data_rows = [
            {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            for row in values[1:]
        ]
        return self._tabular_summary(path, "xlsx", data_rows, headers)

    def _summarize_yaml(self, path: Path) -> dict[str, Any]:
        payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        keys = list(payload.keys()) if isinstance(payload, dict) else []
        return {
            "path": str(path),
            "format": "yaml",
            "rows": len(payload) if isinstance(payload, list) else None,
            "columns": keys,
            "units_detected": self._detect_units(keys),
            "missing_values": None,
            "observable_candidates": self._detect_observables(keys),
            "critical_metadata_missing": not bool(keys),
            "warnings": ["Units are not guaranteed unless explicitly encoded in the file."] if keys else ["No mapping keys found."],
        }

    def _summarize_json(self, path: Path) -> dict[str, Any]:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list) and payload and isinstance(payload[0], dict):
            columns = list(payload[0].keys())
            return self._tabular_summary(path, "json", payload, columns)
        if isinstance(payload, dict):
            columns = list(payload.keys())
            return {
                "path": str(path),
                "format": "json",
                "rows": None,
                "columns": columns,
                "units_detected": self._detect_units(columns),
                "missing_values": None,
                "observable_candidates": self._detect_observables(columns),
                "critical_metadata_missing": not bool(columns),
                "warnings": ["JSON object was treated as metadata, not tabular data."],
            }
        return {
            "path": str(path),
            "format": "json",
            "rows": None,
            "columns": [],
            "units_detected": [],
            "missing_values": None,
            "observable_candidates": [],
            "critical_metadata_missing": True,
            "warnings": ["Unsupported JSON structure for experimental summary."],
        }

    def _summarize_txt(self, path: Path) -> dict[str, Any]:
        lines = path.read_text(encoding="utf-8").splitlines()
        columns: list[str] = []
        if lines and ("," in lines[0] or "\t" in lines[0]):
            delimiter = "," if "," in lines[0] else "\t"
            columns = [item.strip() for item in lines[0].split(delimiter)]
        warnings = ["Plain-text file parsed heuristically; units and semantics may require human review."]
        return {
            "path": str(path),
            "format": "txt",
            "rows": max(0, len(lines) - 1),
            "columns": columns,
            "units_detected": self._detect_units(columns),
            "missing_values": None,
            "observable_candidates": self._detect_observables(columns),
            "critical_metadata_missing": not bool(columns),
            "warnings": warnings,
        }

    def _tabular_summary(
        self,
        path: Path,
        fmt: str,
        rows: list[dict[str, Any]],
        columns: list[str],
    ) -> dict[str, Any]:
        units = self._detect_units(columns)
        missing_values = 0
        for row in rows:
            missing_values += sum(1 for value in row.values() if value in {"", None})
        numeric_columns = self._numeric_column_candidates(rows, columns)
        critical_metadata_missing = not units or not self._has_axis_metadata(columns)
        warnings: list[str] = []
        if not units:
            warnings.append("Units were not detected from column headers.")
        if not self._has_axis_metadata(columns):
            warnings.append("Independent variable / axis metadata could not be inferred from column headers.")
        return {
            "path": str(path),
            "format": fmt,
            "rows": len(rows),
            "columns": columns,
            "units_detected": units,
            "missing_values": missing_values,
            "numeric_columns": numeric_columns,
            "observable_candidates": self._detect_observables(columns),
            "critical_metadata_missing": critical_metadata_missing,
            "warnings": warnings,
        }

    @staticmethod
    def _detect_units(columns: list[str]) -> list[str]:
        units: list[str] = []
        for column in columns:
            lowered = column.lower()
            if "(" in lowered and ")" in lowered:
                units.append(lowered.split("(")[-1].split(")")[0].strip())
            elif "[" in lowered and "]" in lowered:
                units.append(lowered.split("[")[-1].split("]")[0].strip())
            elif "mpa" in lowered:
                units.append("MPa")
            elif "%" in lowered:
                units.append("%")
        return sorted(set(unit for unit in units if unit))

    @staticmethod
    def _detect_observables(columns: list[str]) -> list[str]:
        observables: list[str] = []
        for column in columns:
            lowered = column.lower()
            if "stress" in lowered:
                observables.append("stress")
            if "strain" in lowered:
                observables.append("strain")
            if "texture" in lowered or "orientation" in lowered:
                observables.append("texture")
            if "hardness" in lowered:
                observables.append("hardness")
        return sorted(set(observables))

    @staticmethod
    def _has_axis_metadata(columns: list[str]) -> bool:
        return any(
            token in column.lower()
            for column in columns
            for token in ("time", "strain", "stress", "step", "temperature", "displacement")
        )

    @staticmethod
    def _numeric_column_candidates(rows: list[dict[str, Any]], columns: list[str]) -> list[str]:
        numeric: list[str] = []
        sample_rows = rows[:10]
        for column in columns:
            values = [row.get(column) for row in sample_rows if column in row]
            if values and all(ExperimentalDataAgent._is_number(value) for value in values if value not in {"", None}):
                numeric.append(column)
        return numeric

    @staticmethod
    def _is_number(value: Any) -> bool:
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False

    def _guess_semantic_columns(self, datasets: list[dict[str, Any]]) -> dict[str, str]:
        guesses: dict[str, str] = {}
        for dataset in datasets:
            for column in dataset.get("columns", []):
                lowered = str(column).lower()
                if "true strain" in lowered:
                    guesses[column] = "true_strain"
                elif "engineering strain" in lowered:
                    guesses[column] = "engineering_strain"
                elif "true stress" in lowered:
                    guesses[column] = "true_stress"
                elif "engineering stress" in lowered:
                    guesses[column] = "engineering_stress"
                elif "stress" in lowered:
                    guesses[column] = "possible_stress"
                elif "strain" in lowered:
                    guesses[column] = "possible_strain"
                elif "texture" in lowered or "orientation" in lowered:
                    guesses[column] = "texture_metric"
        return guesses

    def _build_metadata_questions(self, datasets: list[dict[str, Any]]) -> list[str]:
        questions: list[str] = []
        for dataset in datasets:
            columns = dataset.get("columns", [])
            if not dataset.get("units_detected"):
                questions.append(f"Confirm units for dataset {dataset.get('path')}.")
            if any("stress" in str(column).lower() for column in columns) and not any(
                token in " ".join(str(column).lower() for column in columns)
                for token in ("true stress", "engineering stress", "cauchy", "piola")
            ):
                questions.append(f"Clarify the stress definition in {dataset.get('path')}.")
            if any("strain" in str(column).lower() for column in columns) and not any(
                token in " ".join(str(column).lower() for column in columns)
                for token in ("true strain", "engineering strain", "log strain")
            ):
                questions.append(f"Clarify the strain definition in {dataset.get('path')}.")
        deduped: list[str] = []
        for item in questions:
            if item not in deduped:
                deduped.append(item)
        return deduped

    def _build_summary(self, datasets: list[dict[str, Any]], critical_metadata_missing: bool) -> str:
        file_count = len(datasets)
        rows_known = [item["rows"] for item in datasets if item.get("rows") is not None]
        row_summary = f"{sum(rows_known)} total rows across {file_count} file(s)" if rows_known else f"{file_count} file(s)"
        if critical_metadata_missing:
            return f"Loaded experimental data summary for {row_summary}, but critical metadata is still missing."
        return f"Loaded experimental data summary for {row_summary} with sufficient metadata for preliminary alignment."
