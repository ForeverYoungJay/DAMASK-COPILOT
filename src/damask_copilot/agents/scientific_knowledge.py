"""Scientific knowledge aggregation agent for the v1 workflow."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import yaml

from damask_copilot.graph.state import ResearchState
from damask_copilot.mcp_clients.literature_mcp_client import LiteratureMCPClient
from damask_copilot.memory.parameter_store import ParameterStore
from damask_copilot.memory.scientific_memory import ScientificMemoryLayer

DEFAULT_PROJECTS_ROOT = Path("projects")
DEFAULT_FAKE_MATERIALS_DATASET_DIR = Path("data/materials")


class ScientificKnowledgeAgent:
    """Aggregate literature, experimental data, priors, DAMASK docs, and historical records."""

    name = "scientific_knowledge"

    def __init__(
        self,
        *,
        use_llm: bool = False,
        model_name: str | None = None,
        llm_runner=None,
        parameter_store: ParameterStore | None = None,
        scientific_memory: ScientificMemoryLayer | None = None,
        literature_client: LiteratureMCPClient | None = None,
        workspace_root: str = "workspaces",
        projects_root: Path | str = DEFAULT_PROJECTS_ROOT,
        demo_materials_dir: Path | str = DEFAULT_FAKE_MATERIALS_DATASET_DIR,
    ) -> None:
        self.use_llm = use_llm
        self.model_name = model_name
        self.llm_runner = llm_runner
        self.literature_client = literature_client or LiteratureMCPClient()
        self.projects_root = Path(projects_root)
        self.demo_materials_dir = Path(demo_materials_dir)
        self.parameter_store = parameter_store or ParameterStore(data_dir=self.demo_materials_dir)
        if not self.parameter_store.list_ids():
            self.parameter_store.load_library()
        self.workspace_root = Path(workspace_root)
        self.scientific_memory = scientific_memory or ScientificMemoryLayer(
            materials_dir=self.demo_materials_dir,
            workspace_root=self.workspace_root,
            parameter_store=self.parameter_store,
        )

    def run(self, state: ResearchState) -> ResearchState:
        self._hydrate_project_inputs(state)
        state.literature_summary = self._gather_literature(state)
        state.experimental_data = self._gather_experimental_data(state)
        state.known_parameters = self._gather_material_and_parameter_knowledge(state)
        state.damask_capabilities = self._gather_damask_and_history_knowledge(state)
        self.scientific_memory.remember_knowledge_context(state)
        return state.append_trace(
            self.name,
            "knowledge_compiled",
            {
                "material_system": state.material_system,
                "literature_status": state.literature_summary.get("status"),
                "experimental_status": state.experimental_data.get("status"),
                "parameter_source": state.known_parameters.get("source"),
                "previous_simulation_records": len(state.damask_capabilities.get("previous_simulation_records", [])),
            },
        )

    def _hydrate_project_inputs(self, state: ResearchState) -> None:
        project_dir = self._resolve_project_dir(state)
        if project_dir is None:
            return
        state.project_dir = str(project_dir)
        if state.project_name is None and project_dir.parent == self.projects_root:
            state.project_name = project_dir.name
        self._current_goal_text = state.user_goal
        self._current_project_paths = list(state.user_files) + list(state.literature_files) + list(state.experimental_files)
        discovered = self._discover_project_files(project_dir)
        state.user_files = _merge_unique(state.user_files, discovered["user_files"])
        state.literature_files = _merge_unique(state.literature_files, discovered["literature_files"])
        state.experimental_files = _merge_unique(state.experimental_files, discovered["experimental_files"])
        state.source_list_files = _merge_unique(state.source_list_files, discovered["source_list_files"])
        state.literature_sources = _merge_unique_any(
            state.literature_sources,
            self._expand_source_list_files(state.source_list_files),
        )

    def _resolve_project_dir(self, state: ResearchState) -> Path | None:
        if state.project_dir:
            explicit = Path(state.project_dir)
            return explicit if explicit.exists() else None

        if state.project_name:
            named = self.projects_root / state.project_name
            if named.exists():
                return named

        return self._infer_project_dir_from_projects_root(state)

    def _infer_project_dir_from_projects_root(self, state: ResearchState) -> Path | None:
        if not self.projects_root.exists():
            return None

        exact_names = [
            f"{state.material_system}_{state.workflow_type}" if state.material_system and state.workflow_type else None,
            state.material_system,
            self._slugify(state.user_goal),
        ]
        for name in exact_names:
            if not name:
                continue
            candidate = self.projects_root / str(name)
            if candidate.exists():
                return candidate

        candidates = [path for path in self.projects_root.iterdir() if path.is_dir()]
        material_token = (state.material_system or "").lower().strip()
        workflow_token = (state.workflow_type or "").lower().strip()
        if material_token:
            filtered = [path for path in candidates if material_token in path.name.lower()]
            if len(filtered) == 1:
                return filtered[0]
            if workflow_token:
                narrowed = [path for path in filtered if workflow_token in path.name.lower()]
                if len(narrowed) == 1:
                    return narrowed[0]
        if len(candidates) == 1:
            return candidates[0]
        return None

    def _gather_literature(self, state: ResearchState) -> dict[str, Any]:
        if not state.needs_literature and not state.literature_sources and not state.literature_files:
            return {
                "status": "skipped",
                "summary": "Literature retrieval was not required for this workflow.",
                "mechanisms": [],
                "planning_implications": [],
                "experimental_conditions": [],
                "evidence_gaps": [],
                "sources": [],
                "paper_parser_records": [],
                "providers_succeeded": [],
                "planning_evidence": {},
                "parameter_evidence": {},
                "local_file_insights": [],
            }

        external = self._retrieve_literature_external(state)
        local = self._read_local_literature_files(state)
        merged = self._merge_external_and_local(external, local)
        analysis = self._build_deterministic_literature_analysis(state, merged)
        planning_evidence = self._planning_evidence_from_literature(analysis, merged)
        parameter_evidence = self._parameter_evidence_from_literature(state, merged)
        return {
            "status": "collected" if merged.get("used_external_retrieval") or merged.get("used_local_files") else "not_provided",
            "summary": analysis.get("summary") or merged.get("summary") or "No literature summary was available.",
            "mechanisms": list(analysis.get("relevant_mechanisms", [])),
            "candidate_constitutive_models": list(analysis.get("candidate_constitutive_models", [])),
            "planning_implications": list(analysis.get("planning_implications", [])),
            "experimental_conditions": list(analysis.get("experimental_conditions", [])),
            "observables_for_validation": list(analysis.get("observables_for_validation", [])),
            "evidence_gaps": list(analysis.get("evidence_gaps", [])) or list(merged.get("uncertainties", [])),
            "sources": list(merged.get("resolved_sources", [])) or list(state.literature_sources),
            "paper_parser_records": list(merged.get("evidence_items", [])),
            "providers_succeeded": list(merged.get("providers_succeeded", [])),
            "local_files": list(merged.get("local_files", [])),
            "planning_evidence": planning_evidence,
            "parameter_evidence": parameter_evidence,
            "local_file_insights": self._local_literature_file_insights(list(merged.get("local_files", []))),
        }

    def _gather_experimental_data(self, state: ResearchState) -> dict[str, Any]:
        experimental_files = list(state.experimental_files) or self._infer_experimental_files(state)
        if not state.needs_experimental_data and not experimental_files:
            return {
                "status": "skipped",
                "summary": "Experimental-data parsing was not required for this workflow.",
                "datasets": [],
                "curve": None,
                "experimental_conditions": [],
                "observable_candidates": [],
            }

        if not experimental_files:
            return {
                "status": "experimental_data_missing",
                "summary": "No experimental datasets were supplied. This prevents quantitative validation.",
                "datasets": [],
                "curve": None,
                "experimental_conditions": [],
                "observable_candidates": [],
                "critical_metadata_missing": False,
                "needs_human_correction": False,
                "semantic_column_guesses": {},
                "metadata_questions": [],
                "interpretation_summary": "No experimental datasets were supplied; experiment-driven validation was not possible.",
            }

        datasets: list[dict[str, Any]] = []
        observable_candidates: list[str] = []
        critical_metadata_missing = False
        for file_path in experimental_files:
            dataset = self._summarize_experimental_file(Path(file_path))
            datasets.append(dataset)
            for item in dataset.get("observable_candidates", []):
                if item not in observable_candidates:
                    observable_candidates.append(item)
            if dataset.get("critical_metadata_missing", False):
                critical_metadata_missing = True

        summary = {
            "status": "experimental_data_loaded",
            "summary": self._build_experimental_summary(datasets, critical_metadata_missing),
            "datasets": datasets,
            "observable_candidates": observable_candidates,
            "critical_metadata_missing": critical_metadata_missing,
            "needs_human_correction": critical_metadata_missing,
            "semantic_column_guesses": self._guess_semantic_columns(datasets),
            "metadata_questions": self._build_metadata_questions(datasets),
            "interpretation_summary": "Deterministic experimental-data summary completed.",
        }
        summary["curve"] = self._extract_curve(summary)
        summary["experimental_conditions"] = self._infer_experimental_conditions(summary)
        return summary

    def _gather_material_and_parameter_knowledge(self, state: ResearchState) -> dict[str, Any]:
        memory_context = self.scientific_memory.collect_context(
            material_system=state.material_system,
            workflow_type=state.workflow_type,
        )
        card = self.parameter_store.resolve(state.material_system or "generic_material")
        if card is None:
            return {
                "status": "not_found",
                "source": "heuristic_default",
                "payload": {
                    "elastic": {"type": "Hooke", "C_11": 168400000000.0, "C_12": 121400000000.0, "C_44": 75400000000.0},
                    "plastic": {"type": "phenopowerlaw", "n_sl": 20, "h_0_sl-sl": 355000000.0},
                },
                "reported_cp_parameters": {},
                "elastic_constants": {},
                "slip_systems": [],
                "twin_systems": [],
                "phase_information": {},
                "materials_knowledge_graph_hits": list(memory_context.get("materials_knowledge_graph", {}).get("nodes", [])),
            }

        payload = dict(card.parameters or {})
        damask_materialpoint = dict(payload.get("damask", {}).get("materialpoint", {}))
        phase_mapping = dict(damask_materialpoint.get("phase", {}))
        phase_name = next(iter(phase_mapping.keys()), card.material_name)
        phase_payload = dict(phase_mapping.get(phase_name, {}))
        plastic = dict(payload.get("plastic", phase_payload.get("mechanical", {}).get("plastic", {})))
        elastic = dict(payload.get("elastic", phase_payload.get("mechanical", {}).get("elastic", {})))
        phase_info = {
            "material_name": card.material_name,
            "crystal_structure": card.crystal_structure,
            "phase_type": card.phase_type,
            "phase_name": phase_name,
            "lattice": payload.get("damask_lattice") or phase_payload.get("lattice"),
            "description": payload.get("description"),
            "assumptions": list(card.explicit_assumptions),
        }
        return {
            "status": "loaded",
            "source": card.source_path,
            "material_id": card.material_id,
            "material_name": card.material_name,
            "confidence": card.confidence,
            "payload": payload,
            "reported_cp_parameters": plastic,
            "elastic_constants": elastic,
            "slip_systems": self._extract_slip_systems(plastic),
            "twin_systems": self._extract_twin_systems(phase_payload),
            "phase_information": phase_info,
            "materials_knowledge_graph_hits": self._materials_knowledge_graph_hits(card.material_id),
            "scientific_memory_context": memory_context,
            "material_card_summary": {
                "material_id": card.material_id,
                "material_name": card.material_name,
                "crystal_structure": card.crystal_structure,
                "phase_type": card.phase_type,
                "confidence": card.confidence,
            },
        }

    def _gather_damask_and_history_knowledge(self, state: ResearchState) -> dict[str, Any]:
        memory_context = self.scientific_memory.collect_context(
            material_system=state.material_system,
            workflow_type=state.workflow_type,
        )
        return {
            "documentation_sources": self._damask_documentation_search(state),
            "preprocess_tools": [
                "build_material_yaml",
                "build_load_yaml",
                "build_numerics_yaml",
                "build_grid_geometry",
                "validate_damask_inputs",
            ],
            "execution_tools": ["run_damask_grid", "parse_damask_log"],
            "postprocess_tools": [
                "extract_stress_strain",
                "compute_yield_stress",
                "compute_hardening_rate",
                "compare_experiment_simulation",
            ],
            "previous_simulation_records": self._previous_simulation_records(state),
            "shared_memory_layer": memory_context,
        }

    def _retrieve_literature_external(self, state: ResearchState) -> dict[str, Any]:
        try:
            if state.literature_sources:
                return self.literature_client.collect_literature(
                    user_query=state.user_goal,
                    literature_sources=list(state.literature_sources),
                )
            return self.literature_client.search_related_literature(user_query=state.user_goal)
        except Exception as exc:
            return {
                "used_external_retrieval": False,
                "providers_attempted": [],
                "providers_succeeded": [],
                "notes": [],
                "summary": "Literature MCP retrieval failed.",
                "evidence_items": [],
                "resolved_sources": [],
                "uncertainties": [f"{type(exc).__name__}: {exc}"],
                "local_files": [],
                "used_local_files": False,
            }

    def _discover_project_files(self, project_dir: Path) -> dict[str, list[str]]:
        if not project_dir.exists():
            return {
                "user_files": [],
                "literature_files": [],
                "experimental_files": [],
                "source_list_files": [],
            }
        user_files: list[str] = []
        literature_files: list[str] = []
        experimental_files: list[str] = []
        source_list_files: list[str] = []
        ignored_dirs = {"agent_records", "results", ".git", "__pycache__"}
        selected_cases = self._selected_project_cases(project_dir)
        for path in sorted(project_dir.rglob("*")):
            if not path.is_file():
                continue
            if any(part in ignored_dirs for part in path.parts):
                continue
            if path.name.startswith(".") or path.name == ".gitkeep":
                continue
            if path.name in {"research_report.md", "run.log"}:
                continue
            if not self._path_matches_selected_cases(path, project_dir, selected_cases):
                continue
            suffix = path.suffix.lower()
            lowered = str(path).lower()
            if suffix in {".txt", ".md", ".csv"} and any(token in lowered for token in ("source", "doi", "arxiv", "reference_list", "source_list")):
                source_list_files.append(str(path))
            if self._is_literature_file(path):
                literature_files.append(str(path))
            elif self._is_experimental_file(path):
                experimental_files.append(str(path))
            if suffix in {".txt", ".md", ".pdf", ".bib", ".ris", ".yaml", ".yml", ".json", ".csv", ".xlsx", ".xls"}:
                user_files.append(str(path))
        return {
            "user_files": user_files,
            "literature_files": literature_files,
            "experimental_files": experimental_files,
            "source_list_files": source_list_files,
        }

    def _selected_project_cases(self, project_dir: Path) -> set[str]:
        text = self._slugify(self._goal_text_for_case_selection())
        selected: set[str] = set()

        if all(token in text for token in ("single", "tensile")) or "single_crystal" in text:
            selected.add("single_crystal_tensile")
        if "rolling" in text:
            selected.add("cold_rolling_anisotropy")
        if "compression" in text:
            selected.add("uniaxial_compression")
        if "shear" in text:
            selected.add("simple_shear")

        explicit_cases = self._cases_from_existing_paths()
        selected.update(explicit_cases)

        known_cases = self._available_project_cases(project_dir)
        return {case for case in selected if case in known_cases}

    def _goal_text_for_case_selection(self) -> str:
        return getattr(self, "_current_goal_text", "")

    def _cases_from_existing_paths(self) -> set[str]:
        known_case_names = {
            "single_crystal_tensile",
            "cold_rolling_anisotropy",
            "uniaxial_compression",
            "simple_shear",
        }
        selected: set[str] = set()
        for path_text in getattr(self, "_current_project_paths", []):
            parts = {part for part in Path(path_text).parts if part in known_case_names}
            selected.update(parts)
        return selected

    def _available_project_cases(self, project_dir: Path) -> set[str]:
        cases: set[str] = set()
        for category, root_names in {
            "literature": {"pdf", "notes", "bibliographies", "source_lists"},
            "experimental": {"raw", "processed", "metadata"},
        }.items():
            root = project_dir / category
            if not root.exists():
                continue
            for item in root.iterdir():
                if item.is_dir() and item.name not in root_names:
                    cases.add(item.name)
        return cases

    def _path_matches_selected_cases(self, path: Path, project_dir: Path, selected_cases: set[str]) -> bool:
        if not selected_cases:
            return True
        try:
            relative_parts = path.relative_to(project_dir).parts
        except ValueError:
            return True
        for index, part in enumerate(relative_parts):
            if part not in {"literature", "experimental"}:
                continue
            next_index = index + 1
            if next_index >= len(relative_parts):
                return True
            category = part
            candidate = relative_parts[next_index]
            standard_roots = {"literature": {"pdf", "notes", "bibliographies", "source_lists"}, "experimental": {"raw", "processed", "metadata"}}
            if candidate in standard_roots[category]:
                return True
            return candidate in selected_cases
        return True

    @staticmethod
    def _is_literature_file(path: Path) -> bool:
        suffix = path.suffix.lower()
        lowered = str(path).lower()
        if suffix in {".pdf", ".bib", ".ris"}:
            return True
        if suffix in {".md", ".txt"} and "/literature/" in lowered.replace("\\", "/"):
            return True
        return any(token in lowered for token in ("literature", "paper", "article", "reference"))

    @staticmethod
    def _is_experimental_file(path: Path) -> bool:
        suffix = path.suffix.lower()
        lowered = str(path).lower()
        if path.name.startswith("."):
            return False
        if suffix in {".csv", ".xlsx", ".xls", ".txt", ".yaml", ".yml", ".json"}:
            return any(token in lowered for token in ("experiment", "experimental", "tensile", "compression", "stress_strain", "specimen", "metadata", "/raw/", "/processed/"))
        return False

    def _read_local_literature_files(self, state: ResearchState) -> dict[str, Any]:
        notes: list[str] = []
        resolved_files: list[str] = []
        uncertainties: list[str] = []
        for file_path in list(state.literature_files):
            path = Path(file_path)
            if not path.exists():
                uncertainties.append(f"Missing literature file: {file_path}")
                continue
            resolved_files.append(str(path))
            try:
                text = self._extract_local_literature_text(path)
            except Exception as exc:
                uncertainties.append(f"{path.name}: {type(exc).__name__}: {exc}")
                continue
            if text:
                notes.append(f"Local literature file {path.name}: {text[:20000]}")
        return {
            "used_local_files": bool(resolved_files),
            "local_files": resolved_files,
            "notes": notes,
            "uncertainties": uncertainties,
        }

    def _expand_source_list_files(self, source_list_files: list[str]) -> list[str]:
        expanded: list[str] = []
        for file_path in source_list_files:
            path = Path(file_path)
            if not path.exists():
                continue
            for line in path.read_text(encoding="utf-8").splitlines():
                item = line.strip()
                if not item or item.startswith("#") or item in expanded:
                    continue
                expanded.append(item)
        return expanded

    def _merge_external_and_local(self, external: dict[str, Any], local: dict[str, Any]) -> dict[str, Any]:
        merged = dict(external)
        merged["used_local_files"] = bool(local.get("used_local_files"))
        merged["local_files"] = list(local.get("local_files", []))
        merged["notes"] = list(external.get("notes", [])) + [
            item for item in local.get("notes", []) if item not in external.get("notes", [])
        ]
        merged["uncertainties"] = list(external.get("uncertainties", [])) + [
            item for item in local.get("uncertainties", []) if item not in external.get("uncertainties", [])
        ]
        if local.get("used_local_files"):
            summary = str(merged.get("summary") or "").strip()
            merged["summary"] = f"{summary} Local literature files were also supplied.".strip()
        return merged

    def _build_deterministic_literature_analysis(self, state: ResearchState, external: dict[str, Any]) -> dict[str, Any]:
        notes = [str(item) for item in external.get("notes", []) if item]
        text_blob = "\n".join(notes + [state.user_goal])
        lowered = text_blob.lower()

        mechanisms: list[str] = []
        if "slip" in lowered:
            mechanisms.append("slip-mediated plasticity")
        if "twinning" in lowered:
            mechanisms.append("twinning")
        if "hardening" in lowered:
            mechanisms.append("strain hardening")
        if "texture" in lowered or "orientation" in lowered:
            mechanisms.append("texture evolution")

        models: list[str] = []
        if "phenopowerlaw" in lowered:
            models.append("phenopowerlaw crystal plasticity")
        if "crystal plasticity" in lowered or "slip" in lowered:
            models.append("slip-based crystal plasticity")
        if "twinning" in lowered:
            models.append("crystal plasticity with twinning-capable kinematics")

        observables: list[str] = []
        if "stress" in lowered and "strain" in lowered:
            observables.append("stress_strain_curve")
        else:
            if "stress" in lowered:
                observables.append("stress")
            if "strain" in lowered:
                observables.append("strain")
        if "texture" in lowered or "orientation" in lowered:
            observables.append("texture_evolution")
        if "slip" in lowered:
            observables.append("slip_activity_proxy")

        experimental_conditions: list[str] = []
        if "tension" in lowered:
            experimental_conditions.append("uniaxial_tension")
        if "compression" in lowered:
            experimental_conditions.append("uniaxial_compression")
        if "shear" in lowered:
            experimental_conditions.append("simple_shear")
        if "cyclic" in lowered:
            experimental_conditions.append("cyclic_loading")
        if "temperature" in lowered:
            experimental_conditions.append("temperature_condition_reported")
        if "strain rate" in lowered or "strain-rate" in lowered or "rate sensitivity" in lowered:
            experimental_conditions.append("strain_rate_sensitive")

        planning_implications: list[str] = []
        if any("slip" in item for item in mechanisms):
            planning_implications.append("Choose a crystal-plasticity formulation that can represent FCC slip-driven deformation.")
        if "stress_strain_curve" in observables:
            planning_implications.append("Include volume-averaged stress-strain outputs to test the primary literature observable.")
        if "texture_evolution" in observables:
            planning_implications.append("Track orientation or texture-related outputs if the study aims to compare microstructural evolution.")
        if not planning_implications:
            planning_implications.append("Use literature only as qualitative framing and keep the first simulation plan conservative.")

        evidence_gaps: list[str] = []
        if not notes:
            evidence_gaps.append("No external or local literature evidence was available to ground the study design.")
        if "stress_strain_curve" not in observables:
            evidence_gaps.append("The retrieved evidence does not clearly define a primary validation observable.")
        if not models:
            evidence_gaps.append("The retrieved evidence does not clearly support a constitutive-model choice.")

        summary = external.get("summary") or (
            "Literature evidence was reviewed to guide hypothesis generation and conservative simulation planning."
            if notes
            else "No literature evidence was available; proceed with conservative planning assumptions only."
        )
        return {
            "summary": summary,
            "relevant_mechanisms": mechanisms,
            "candidate_constitutive_models": models,
            "experimental_conditions": experimental_conditions,
            "observables_for_validation": observables,
            "planning_implications": planning_implications,
            "evidence_gaps": evidence_gaps,
        }

    def _planning_evidence_from_literature(self, analysis: dict[str, Any], merged: dict[str, Any]) -> dict[str, Any]:
        return {
            "mechanisms": list(analysis.get("relevant_mechanisms", [])),
            "candidate_constitutive_models": list(analysis.get("candidate_constitutive_models", [])),
            "planning_implications": list(analysis.get("planning_implications", [])),
            "observables_for_validation": list(analysis.get("observables_for_validation", [])),
            "experimental_conditions": list(analysis.get("experimental_conditions", [])),
            "evidence_gaps": list(analysis.get("evidence_gaps", [])),
            "providers_succeeded": list(merged.get("providers_succeeded", [])),
            "usable_sources": list(merged.get("resolved_sources", [])),
        }

    def _parameter_evidence_from_literature(self, state: ResearchState, merged: dict[str, Any]) -> dict[str, Any]:
        notes_blob = "\n".join(str(item) for item in merged.get("notes", []) if item).lower()
        candidate_keys = []
        for token in ("xi_0_sl", "xi_inf_sl", "h_0_sl-sl", "n_sl", "c_11", "c_12", "c_44"):
            if token.lower() in notes_blob:
                candidate_keys.append(token)
        return {
            "usable_sources": list(merged.get("resolved_sources", [])),
            "candidate_parameter_keys": candidate_keys,
            "supports_parameter_lookup": bool(candidate_keys or merged.get("resolved_sources")),
            "intended_material_system": state.material_system,
        }

    def _local_literature_file_insights(self, local_files: list[str]) -> list[dict[str, Any]]:
        insights: list[dict[str, Any]] = []
        for file_path in local_files:
            path = Path(file_path)
            lowered = str(path).lower()
            if "note" in lowered:
                role = "planning_note"
                planning_help = "Useful for extracting mechanisms, validation observables, and study framing."
            elif "source_list" in lowered or "bibliograph" in lowered:
                role = "source_index"
                planning_help = "Useful for expanding literature retrieval and tracing planning evidence to references."
            elif path.suffix.lower() == ".pdf":
                role = "primary_paper"
                planning_help = "Useful as primary evidence for both constitutive assumptions and project planning."
            else:
                role = "literature_asset"
                planning_help = "Useful as supporting scientific context."
            insights.append({
                "path": str(path),
                "role": role,
                "planning_help": planning_help,
            })
        return insights

    def _infer_experimental_files(self, state: ResearchState) -> list[str]:
        supported = {".csv", ".xlsx", ".xls", ".yaml", ".yml", ".json", ".txt"}
        return [path for path in state.user_files if Path(path).suffix.lower() in supported]

    def _summarize_experimental_file(self, path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8") as handle:
                reader = csv.DictReader(handle)
                rows = list(reader)
                columns = list(reader.fieldnames or [])
            return self._tabular_summary(path, "csv", rows, columns)
        if suffix in {".yaml", ".yml"}:
            payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            keys = list(payload.keys()) if isinstance(payload, dict) else []
            return {
                "path": str(path),
                "format": "yaml",
                "rows": len(payload) if isinstance(payload, list) else None,
                "columns": keys,
                "units_detected": self._detect_units(keys),
                "missing_values": None,
                "observable_candidates": self._detect_observables(keys),
                "critical_metadata_missing": not bool(keys),
                "warnings": ["Units are not guaranteed unless explicitly encoded in the file."] if keys else ["No mapping keys found."],
            }
        if suffix == ".json":
            payload = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(payload, list) and payload and isinstance(payload[0], dict):
                columns = list(payload[0].keys())
                return self._tabular_summary(path, "json", payload, columns)
            if isinstance(payload, dict):
                columns = list(payload.keys())
                return {
                    "path": str(path),
                    "format": "json",
                    "rows": None,
                    "columns": columns,
                    "units_detected": self._detect_units(columns),
                    "missing_values": None,
                    "observable_candidates": self._detect_observables(columns),
                    "critical_metadata_missing": not bool(columns),
                    "warnings": ["JSON object was treated as metadata, not tabular data."],
                }
        if suffix == ".txt":
            lines = path.read_text(encoding="utf-8").splitlines()
            columns: list[str] = []
            if lines and ("," in lines[0] or "\t" in lines[0]):
                delimiter = "," if "," in lines[0] else "\t"
                columns = [item.strip() for item in lines[0].split(delimiter)]
            return {
                "path": str(path),
                "format": "txt",
                "rows": max(0, len(lines) - 1),
                "columns": columns,
                "units_detected": self._detect_units(columns),
                "missing_values": None,
                "observable_candidates": self._detect_observables(columns),
                "critical_metadata_missing": not bool(columns),
                "warnings": ["Plain-text file parsed heuristically; units and semantics may require human review."],
            }
        if suffix in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return {
                    "path": str(path),
                    "format": "xlsx",
                    "rows": None,
                    "columns": [],
                    "units_detected": [],
                    "missing_values": None,
                    "observable_candidates": [],
                    "critical_metadata_missing": True,
                    "warnings": ["openpyxl is not available; XLSX metadata could not be inspected."],
                }
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            values = list(sheet.iter_rows(values_only=True))
            headers = [str(item) if item is not None else "" for item in (values[0] if values else [])]
            data_rows = [
                {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
                for row in values[1:]
            ]
            return self._tabular_summary(path, "xlsx", data_rows, headers)
        return {
            "path": str(path),
            "format": suffix.lstrip(".") or "unknown",
            "rows": None,
            "columns": [],
            "units_detected": [],
            "missing_values": None,
            "observable_candidates": [],
            "critical_metadata_missing": True,
            "warnings": [f"Unsupported experimental file format: {suffix or 'unknown'}"],
        }

    def _tabular_summary(self, path: Path, fmt: str, rows: list[dict[str, Any]], columns: list[str]) -> dict[str, Any]:
        units = self._detect_units(columns)
        missing_values = 0
        for row in rows:
            missing_values += sum(1 for value in row.values() if value in {"", None})
        numeric_columns = self._numeric_column_candidates(rows, columns)
        critical_metadata_missing = not units or not self._has_axis_metadata(columns)
        warnings: list[str] = []
        if not units:
            warnings.append("Units were not detected from column headers.")
        if not self._has_axis_metadata(columns):
            warnings.append("Independent variable / axis metadata could not be inferred from column headers.")
        preview_rows = rows[:10]
        return {
            "path": str(path),
            "file": path.name,
            "format": fmt,
            "rows": len(rows),
            "columns": columns,
            "units_detected": units,
            "missing_values": missing_values,
            "numeric_columns": numeric_columns,
            "observable_candidates": self._detect_observables(columns),
            "critical_metadata_missing": critical_metadata_missing,
            "warnings": warnings,
            "preview_rows": preview_rows,
        }

    @staticmethod
    def _detect_units(columns: list[str]) -> list[str]:
        units: list[str] = []
        for column in columns:
            lowered = column.lower()
            if "(" in lowered and ")" in lowered:
                units.append(lowered.split("(")[-1].split(")")[0].strip())
            elif "[" in lowered and "]" in lowered:
                units.append(lowered.split("[")[-1].split("]")[0].strip())
            elif "mpa" in lowered:
                units.append("MPa")
            elif "%" in lowered:
                units.append("%")
        return sorted(set(unit for unit in units if unit))

    @staticmethod
    def _detect_observables(columns: list[str]) -> list[str]:
        observables: list[str] = []
        for column in columns:
            lowered = column.lower()
            if "stress" in lowered:
                observables.append("stress")
            if "strain" in lowered:
                observables.append("strain")
            if "texture" in lowered or "orientation" in lowered:
                observables.append("texture")
            if "hardness" in lowered:
                observables.append("hardness")
        return sorted(set(observables))

    @staticmethod
    def _has_axis_metadata(columns: list[str]) -> bool:
        return any(
            token in column.lower()
            for column in columns
            for token in ("time", "strain", "stress", "step", "temperature", "displacement")
        )

    @staticmethod
    def _numeric_column_candidates(rows: list[dict[str, Any]], columns: list[str]) -> list[str]:
        numeric: list[str] = []
        sample_rows = rows[:10]
        for column in columns:
            values = [row.get(column) for row in sample_rows if column in row]
            if values and all(ScientificKnowledgeAgent._is_number(value) for value in values if value not in {"", None}):
                numeric.append(column)
        return numeric

    @staticmethod
    def _is_number(value: Any) -> bool:
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False

    def _guess_semantic_columns(self, datasets: list[dict[str, Any]]) -> dict[str, str]:
        guesses: dict[str, str] = {}
        for dataset in datasets:
            for column in dataset.get("columns", []):
                lowered = str(column).lower()
                if "true strain" in lowered:
                    guesses[column] = "true_strain"
                elif "engineering strain" in lowered:
                    guesses[column] = "engineering_strain"
                elif "true stress" in lowered:
                    guesses[column] = "true_stress"
                elif "engineering stress" in lowered:
                    guesses[column] = "engineering_stress"
                elif "stress" in lowered:
                    guesses[column] = "possible_stress"
                elif "strain" in lowered:
                    guesses[column] = "possible_strain"
                elif "texture" in lowered or "orientation" in lowered:
                    guesses[column] = "texture_metric"
        return guesses

    def _build_metadata_questions(self, datasets: list[dict[str, Any]]) -> list[str]:
        questions: list[str] = []
        for dataset in datasets:
            columns = dataset.get("columns", [])
            if not dataset.get("units_detected"):
                questions.append(f"Confirm units for dataset {dataset.get('path')}.")
            lowered_columns = " ".join(str(column).lower() for column in columns)
            if any("stress" in str(column).lower() for column in columns) and not any(
                token in lowered_columns for token in ("true stress", "engineering stress", "cauchy", "piola")
            ):
                questions.append(f"Clarify the stress definition in {dataset.get('path')}.")
            if any("strain" in str(column).lower() for column in columns) and not any(
                token in lowered_columns for token in ("true strain", "engineering strain", "log strain")
            ):
                questions.append(f"Clarify the strain definition in {dataset.get('path')}.")
        deduped: list[str] = []
        for item in questions:
            if item not in deduped:
                deduped.append(item)
        return deduped

    def _build_experimental_summary(self, datasets: list[dict[str, Any]], critical_metadata_missing: bool) -> str:
        file_count = len(datasets)
        rows_known = [item["rows"] for item in datasets if item.get("rows") is not None]
        row_summary = f"{sum(rows_known)} total rows across {file_count} file(s)" if rows_known else f"{file_count} file(s)"
        if critical_metadata_missing:
            return f"Loaded experimental data summary for {row_summary}, but critical metadata is still missing."
        return f"Loaded experimental data summary for {row_summary} with sufficient metadata for preliminary alignment."

    def _extract_curve(self, summary: dict[str, Any]) -> dict[str, list[float]] | None:
        for dataset in summary.get("datasets", []) or []:
            preview = dataset.get("preview_rows")
            if not isinstance(preview, list):
                continue
            strain: list[float] = []
            stress: list[float] = []
            for row in preview:
                if not isinstance(row, dict):
                    continue
                lowered = {str(key).lower(): value for key, value in row.items()}
                strain_key = next((key for key in lowered if "strain" in key), None)
                stress_key = next((key for key in lowered if "stress" in key), None)
                if strain_key is None or stress_key is None:
                    continue
                try:
                    strain.append(float(lowered[strain_key]))
                    stress.append(float(lowered[stress_key]))
                except (TypeError, ValueError):
                    continue
            if strain and stress:
                return {"strain": strain, "stress": stress}
        return None

    def _infer_experimental_conditions(self, summary: dict[str, Any]) -> list[str]:
        conditions: list[str] = []
        for dataset in summary.get("datasets", []) or []:
            file_name = dataset.get("file") or dataset.get("path")
            if file_name:
                conditions.append(f"dataset={file_name}")
            columns = dataset.get("columns") or []
            if columns:
                conditions.append(f"columns={columns}")
        interpretation = summary.get("interpretation_summary")
        if interpretation:
            conditions.append(str(interpretation))
        return conditions

    def _extract_slip_systems(self, plastic: dict[str, Any]) -> list[dict[str, Any]]:
        if not plastic:
            return []
        slip_count = plastic.get("N_sl")
        if isinstance(slip_count, list):
            total = int(sum(item for item in slip_count if isinstance(item, (int, float))))
        elif isinstance(slip_count, (int, float)):
            total = int(slip_count)
        else:
            total = 0
        return [{
            "family": "slip",
            "count": total,
            "hardening_law": plastic.get("type"),
            "reference_rate": plastic.get("dot_gamma_0_sl"),
        }] if total else []

    def _extract_twin_systems(self, phase_payload: dict[str, Any]) -> list[dict[str, Any]]:
        mechanical = dict(phase_payload.get("mechanical", {}))
        twin = dict(mechanical.get("twin", {}))
        if not twin:
            return []
        return [{"family": "twin", "definition": twin}]

    def _materials_knowledge_graph_hits(self, material_id: str) -> list[dict[str, Any]]:
        index_path = self.demo_materials_dir / "index.yaml"
        if not index_path.exists():
            return []
        payload = yaml.safe_load(index_path.read_text(encoding="utf-8")) or {}
        entry = dict(payload.get("materials", {}).get(material_id, {}))
        if not entry:
            return []
        return [{
            "material_id": material_id,
            "aliases": list(entry.get("aliases", [])),
            "file": entry.get("file"),
            "source": str(index_path),
        }]

    def _damask_documentation_search(self, state: ResearchState) -> list[dict[str, Any]]:
        doc_paths = [
            Path("README.md"),
            Path("docs/preprocessing_api_inventory.md"),
            Path("docs/postprocessing_api_inventory.md"),
            Path("docs/miscellaneous_api_inventory.md"),
        ]
        keywords = [
            state.workflow_type or "",
            state.material_system or "",
            "DAMASK",
            "material.yaml",
            "spectral_basic",
            "phenopowerlaw",
        ]
        hits: list[dict[str, Any]] = []
        for doc_path in doc_paths:
            if not doc_path.exists():
                continue
            text = doc_path.read_text(encoding="utf-8", errors="ignore")
            snippets: list[str] = []
            lowered = text.lower()
            for keyword in keywords:
                token = keyword.lower().strip()
                if not token or token not in lowered:
                    continue
                idx = lowered.find(token)
                start = max(0, idx - 120)
                end = min(len(text), idx + 220)
                snippet = " ".join(text[start:end].split())
                if snippet and snippet not in snippets:
                    snippets.append(snippet)
            hits.append({
                "path": str(doc_path),
                "matched": bool(snippets),
                "snippets": snippets[:3],
            })
        return hits

    def _previous_simulation_records(self, state: ResearchState) -> list[dict[str, Any]]:
        if not self.workspace_root.exists():
            return []
        material_key = (state.material_system or "").lower()
        matches: list[dict[str, Any]] = []
        for state_path in sorted(self.workspace_root.glob("*/research_state.json")):
            try:
                payload = json.loads(state_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            selected_material = str(payload.get("selected_material_key") or "").lower()
            goal_material = str((payload.get("goal") or {}).get("material_system") or "").lower()
            if material_key and material_key not in {selected_material, goal_material}:
                continue
            workspace = state_path.parent
            matches.append({
                "workspace": workspace.name,
                "state_path": str(state_path),
                "report_path": str(workspace / "report.md") if (workspace / "report.md").exists() else str(workspace / "research_report.md") if (workspace / "research_report.md").exists() else None,
                "selected_material_key": payload.get("selected_material_key"),
                "goal": payload.get("goal"),
                "has_run_report": bool(payload.get("run_report")),
                "has_checker_report": bool(payload.get("checker_report")),
            })
        return matches[:10]

    def _extract_local_literature_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".md", ".txt", ".bib", ".ris", ".yaml", ".yml", ".json"}:
            return path.read_text(encoding="utf-8")
        if suffix == ".pdf":
            try:
                from pypdf import PdfReader
            except Exception:
                return f"PDF file available at {path}. Text extraction library is not installed."
            reader = PdfReader(str(path))
            texts: list[str] = []
            total_chars = 0
            for page in reader.pages:
                page_text = page.extract_text() or ""
                if not page_text:
                    continue
                texts.append(page_text)
                total_chars += len(page_text)
                if total_chars >= 40000:
                    break
            return "\n".join(texts).strip() or f"PDF file available at {path}, but no extractable text was found."
        return f"Local literature file available at {path}."

    @staticmethod
    def _slugify(value: str | None) -> str:
        text = (value or "").strip().lower()
        safe = "".join(ch if ch.isalnum() else "_" for ch in text)
        compact = "_".join(part for part in safe.split("_") if part)
        return compact[:80] or "project"


def _merge_unique(existing: list[str], new_items: list[str]) -> list[str]:
    merged = list(existing)
    for item in new_items:
        if item not in merged:
            merged.append(item)
    return merged


def _merge_unique_any(existing: list[Any], new_items: list[Any]) -> list[Any]:
    merged = list(existing)
    for item in new_items:
        if item not in merged:
            merged.append(item)
    return merged
